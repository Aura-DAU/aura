"""
xlsx_parser.py — Production XLSX timetable parser for DAU.

Parses the official DAU Lecture Time-Table Excel file into structured
records suitable for insertion into the ``timetable_master`` PostgreSQL
table.  Handles merged cells, multi-section courses, continuation rows
(elective options listed below the batch header), and both the DAU-specific
6-column-per-day layout (Winter 2026) and the 3-column-per-day layout
(Autumn 2026).

Architecture
============
The DAU timetable XLSX has this layout on the "Time-Table" sheet:

    Row 1-3 : Title / header rows (ignored)
    Row 4   : Column headers — Time | Batch | Monday | Tuesday | Wednesday | Thursday | Friday
    Row 6+  : Time-slot blocks

Each **time-slot block** starts with a row where column A contains a time
range (e.g. "8:00 - 8:50").  Below it, each row describes one batch's
classes for that slot across the 5 weekdays.

Each **day** occupies 6 contiguous columns:
    [CourseCode, CourseName, Credits(L-T-P-C), CourseType, FacultyInitials, Room]

Column mapping (1-indexed):
    Monday    : D(4)  - I(9)
    Tuesday   : K(11) - P(16)
    Wednesday : R(18) - W(23)
    Thursday  : Y(25) - AD(30)
    Friday    : AF(32)- AK(37)

Sections are encoded inside the course name: "Data Structures (Sec A)"
→ section = "A", clean name = "Data Structures".

Usage
=====
    from pipeline.timetable.xlsx_parser import parse_timetable
    records = parse_timetable("path/to/Lecture_Time_Table.xlsx")
    # records is a list of TimetableRecord dataclasses

CLI:
    python -m pipeline.timetable.xlsx_parser path/to/file.xlsx [--json] [--csv]
"""

from __future__ import annotations

import json
import logging
import re
import sys
from dataclasses import asdict, dataclass, field
from pathlib import Path
from typing import Optional

logger = logging.getLogger("aura.timetable.xlsx_parser")

# ── Day → column offsets (1-indexed) ──────────────────────────────────────
# V1 format (Winter 2026): Each day has 6 columns: code, name, credits, type, faculty, room
DAY_COLUMNS_V1: dict[str, tuple[int, ...]] = {
    "Monday":    (4, 5, 6, 7, 8, 9),
    "Tuesday":   (11, 12, 13, 14, 15, 16),
    "Wednesday": (18, 19, 20, 21, 22, 23),
    "Thursday":  (25, 26, 27, 28, 29, 30),
    "Friday":    (32, 33, 34, 35, 36, 37),
}
# Backward compat alias
DAY_COLUMNS = DAY_COLUMNS_V1

# V2 format (Autumn 2026): Each day has 3 columns: course_code, faculty, room
# Days start at C4, C8, C12, C16, C20 (col gap of 4, but only 3 data cols)
DAY_COLUMNS_V2: dict[str, tuple[int, ...]] = {
    "Monday":    (4, 5, 6),
    "Tuesday":   (8, 9, 10),
    "Wednesday": (12, 13, 14),
    "Thursday":  (16, 17, 18),
    "Friday":    (20, 21, 22),
}

DAY_OF_WEEK: dict[str, int] = {
    "Monday": 0, "Tuesday": 1, "Wednesday": 2,
    "Thursday": 3, "Friday": 4,
}

# ── Time-slot detection regex ─────────────────────────────────────────────
# Matches patterns like "8:00 - 8:50", "14:00 - 14:50"
TIME_SLOT_RE = re.compile(
    r"(\d{1,2}:\d{2})\s*[-–]\s*(\d{1,2}:\d{2})"
)

# ── Section extraction regex ──────────────────────────────────────────────
# Matches "(Sec A)", "(Sec B)", "(Section A)", case-insensitive
SECTION_RE = re.compile(
    r"\s*\((?:Sec(?:tion)?)\s+([A-Z0-9]+)\)", re.IGNORECASE
)

# ── Batch parsing regex ───────────────────────────────────────────────────
# "BTech Sem-II (ICT + CS)" → program=BTech, semester=II, branch=ICT + CS
# "MTech Sem-II (ICT-ML)"   → program=MTech, semester=II, branch=ICT-ML
# "MSc Sem-II (IT)"         → program=MSc, semester=II, branch=IT
BATCH_RE = re.compile(
    r"(BTech|MTech|MSc|PhD)\s+Sem[- ]?([IVXLC]+|\d+)\s*"
    r"(?:\(([^)]+)\))?",
    re.IGNORECASE,
)

# V2 batch regex for "Btech 1st Yr", "Btech 2nd Year", "Btech 3rd Yr (Core)"
# Also handles typos like "Btech 3r Yr" (missing 'd')
BATCH_V2_YEAR_RE = re.compile(
    r"(BTech|Btech|BTECH)\s+(\d+)(?:st|nd|rd?|th)?\s+Y(?:ea)?r",
    re.IGNORECASE,
)
# Fallback: "Btech Core" without a year
BATCH_V2_CORE_RE = re.compile(
    r"(BTech|Btech|BTECH)\s+(?:\(?\s*Core\s*\)?)",
    re.IGNORECASE,
)

# V2 section regex: extract "(A)" or "( C )" from course code like "IC101 (A)"
CODE_SECTION_RE = re.compile(
    r"\s*\(\s*([A-Z0-9]+)\s*\)\s*$", re.IGNORECASE
)

# ── Roman numeral → int ──────────────────────────────────────────────────
ROMAN_MAP = {
    "I": 1, "II": 2, "III": 3, "IV": 4, "V": 5,
    "VI": 6, "VII": 7, "VIII": 8, "IX": 9, "X": 10,
}


@dataclass
class TimetableRecord:
    """One class session in the master timetable."""

    # ── Batch identity ────────────────────────────────────────────────
    batch_raw: str              # Original string, e.g. "BTech Sem-II (ICT + CS)"
    program: str                # "BTech", "MTech", "MSc"
    branch: str                 # "ICT + CS", "EVD", "MnC", "ICT-ML", etc.
    semester: int               # 2, 4, 6, 8 ...
    year: int                   # Derived: ceil(semester / 2)

    # ── Schedule ──────────────────────────────────────────────────────
    day: str                    # "Monday" .. "Friday"
    day_of_week: int            # 0=Monday .. 4=Friday
    start_time: str             # "08:00" (zero-padded HH:MM)
    end_time: str               # "08:50"

    # ── Course ────────────────────────────────────────────────────────
    course_code: str            # "IT205"
    course_name: str            # "Data Structures" (section stripped)
    section: Optional[str]      # "A", "B", or None
    credits: str                # "3-0-2-4" (L-T-P-C)
    course_type: str            # "Core", "ICT & Technical Elective", etc.
    session_type: str           # "lecture", "lab", "tutorial" (inferred)

    # ── People & place ────────────────────────────────────────────────
    faculty_initials: str       # "PD", "AR2", "SS/VS"
    room: str                   # "LT-1", "CEP110"


def _zero_pad_time(t: str) -> str:
    """Ensure time is in HH:MM format: '8:00' → '08:00'."""
    parts = t.strip().split(":")
    if len(parts) == 2:
        return f"{int(parts[0]):02d}:{parts[1]}"
    return t.strip()


def _parse_roman(s: str) -> int:
    """Convert a Roman numeral string to int. Falls back to int()."""
    s = s.strip().upper()
    if s in ROMAN_MAP:
        return ROMAN_MAP[s]
    try:
        return int(s)
    except ValueError:
        logger.warning("Could not parse semester number: %r, defaulting to 0", s)
        return 0


def _parse_batch(batch_str: str) -> dict:
    """
    Parse a batch string into structured components.

    >>> _parse_batch("BTech Sem-II (ICT + CS)")
    {'program': 'BTech', 'branch': 'ICT + CS', 'semester': 2, 'year': 1}
    """
    m = BATCH_RE.search(batch_str)
    if not m:
        logger.warning("Could not parse batch string: %r", batch_str)
        return {
            "program": batch_str.strip(),
            "branch": "",
            "semester": 0,
            "year": 0,
        }
    program = m.group(1)
    semester = _parse_roman(m.group(2))
    branch = (m.group(3) or "").strip()
    year = (semester + 1) // 2

    return {
        "program": program,
        "branch": branch,
        "semester": semester,
        "year": year,
    }


def _extract_section(course_name: str) -> tuple[str, Optional[str]]:
    """
    Extract section identifier from a course name.

    Returns (clean_name, section_or_None).

    >>> _extract_section("Data Structures (Sec A)")
    ('Data Structures', 'A')
    >>> _extract_section("Introduction to Robotics (ICT only)")
    ('Introduction to Robotics (ICT only)', None)
    """
    m = SECTION_RE.search(course_name)
    if m:
        clean = course_name[:m.start()].strip()
        section = m.group(1).upper()
        return clean, section
    return course_name.strip(), None


def _infer_session_type(course_name: str, credits: str) -> str:
    """
    Infer the session type from the course name and credits string.

    Credits format: L-T-P-C (Lecture-Tutorial-Practical-Credits)
    - "3-0-0-3" → lecture
    - "1-0-2-2" → lab (course name likely has "Lab")
    - "0-1-0-1" → tutorial

    Heuristic priority:
    1. If "lab" appears in the course name → "lab"
    2. If "tutorial" appears in the course name → "tutorial"
    3. If L=0 and P>0 → "lab"
    4. If L=0 and T>0 → "tutorial"
    5. Default → "lecture"
    """
    name_lower = course_name.lower()
    if "lab" in name_lower:
        return "lab"
    if "tutorial" in name_lower:
        return "tutorial"

    parts = credits.split("-")
    if len(parts) >= 3:
        try:
            l, t, p = int(parts[0]), int(parts[1]), int(parts[2])
            if l == 0 and p > 0:
                return "lab"
            if l == 0 and t > 0 and p == 0:
                return "tutorial"
        except ValueError:
            pass

    return "lecture"


def _cell_str(ws, row: int, col: int) -> str:
    """Read a cell value as a stripped string, returning '' for None."""
    v = ws.cell(row=row, column=col).value
    if v is None:
        return ""
    return str(v).strip()


def parse_timetable(
    xlsx_path: str | Path,
    sheet_name: str = "Time-Table",
) -> list[TimetableRecord]:
    """
    Parse a DAU Lecture Time-Table XLSX file into a list of
    :class:`TimetableRecord` objects.

    Auto-detects the format:
    - V1 (Winter 2026): 6 cols/day, sheet "Time-Table", batch in col B
    - V2 (Autumn 2026): 3 cols/day, sheet "Lecture (Update)", batch in col C3

    Parameters
    ----------
    xlsx_path : str or Path
        Path to the XLSX file.
    sheet_name : str
        Name of the sheet containing the timetable.
        Use 'auto' to auto-detect from available sheets.

    Returns
    -------
    list[TimetableRecord]
        One record per (batch, day, time-slot, course) combination.
    """
    import openpyxl

    xlsx_path = Path(xlsx_path)
    if not xlsx_path.exists():
        raise FileNotFoundError(f"Timetable file not found: {xlsx_path}")

    # Auto-detect sheet and format
    wb = openpyxl.load_workbook(str(xlsx_path), data_only=True, read_only=True)
    available_sheets = wb.sheetnames
    wb.close()

    if sheet_name == "auto" or sheet_name not in available_sheets:
        # Try known sheet names in priority order
        for candidate in ["Time-Table", "Lecture (Update)"]:
            if candidate in available_sheets:
                sheet_name = candidate
                break
        else:
            # Use first sheet as fallback
            sheet_name = available_sheets[0]

    # Detect format based on sheet name and column structure
    if sheet_name == "Lecture (Update)":
        logger.info("Detected V2 (Autumn 2026) format, sheet: %s", sheet_name)
        return _parse_v2_3col(xlsx_path, sheet_name)

    # V1 format (original)
    logger.info("Detected V1 (Winter 2026) format, sheet: %s", sheet_name)

    wb = openpyxl.load_workbook(str(xlsx_path), data_only=True)

    if sheet_name not in wb.sheetnames:
        raise ValueError(
            f"Sheet '{sheet_name}' not found. Available: {wb.sheetnames}"
        )

    ws = wb[sheet_name]
    max_row = ws.max_row
    logger.info("Sheet '%s': %d rows x %d cols", sheet_name, max_row, ws.max_column)

    # ── Step 1: Discover time-slot boundaries ─────────────────────────
    # A time-slot row is any row where column A matches the time pattern.
    time_slots: list[dict] = []
    for row_idx in range(1, max_row + 1):
        a_val = _cell_str(ws, row_idx, 1)
        m = TIME_SLOT_RE.match(a_val)
        if m:
            time_slots.append({
                "row": row_idx,
                "start": _zero_pad_time(m.group(1)),
                "end": _zero_pad_time(m.group(2)),
                "raw": a_val,
            })

    if not time_slots:
        raise ValueError("No time-slot rows found in column A. Is this the right file?")

    logger.info("Found %d time slots: %s", len(time_slots), [s["raw"] for s in time_slots])

    # ── Step 2: For each time-slot block, extract classes ─────────────
    records: list[TimetableRecord] = []

    for slot_idx, slot in enumerate(time_slots):
        slot_start_row = slot["row"]
        # Block ends at the row before the next time-slot, or at max_row
        if slot_idx + 1 < len(time_slots):
            slot_end_row = time_slots[slot_idx + 1]["row"] - 1
        else:
            slot_end_row = max_row

        current_batch_raw = ""
        current_batch_parsed: dict = {}

        # Iterate every row in this time-slot block
        for row_idx in range(slot_start_row, slot_end_row + 1):
            # Check if column B has a batch name (new batch or continuation)
            b_val = _cell_str(ws, row_idx, 2)
            if b_val:
                current_batch_raw = b_val
                current_batch_parsed = _parse_batch(b_val)

            # Skip rows with no batch context (header/title rows)
            if not current_batch_raw:
                continue

            # ── Extract classes for each day ──────────────────────────
            for day_name, cols in DAY_COLUMNS.items():
                code_col, name_col, credits_col, type_col, faculty_col, room_col = cols

                course_code = _cell_str(ws, row_idx, code_col)
                course_name_raw = _cell_str(ws, row_idx, name_col)

                # Skip empty cells and slot-label cells (e.g. "Slot-1")
                if not course_code or course_code.startswith("Slot"):
                    continue
                if not course_name_raw:
                    continue

                # Parse the remaining columns
                credits = _cell_str(ws, row_idx, credits_col)
                course_type = _cell_str(ws, row_idx, type_col)
                faculty = _cell_str(ws, row_idx, faculty_col)
                room = _cell_str(ws, row_idx, room_col)

                # Extract section from course name
                course_name_clean, section = _extract_section(course_name_raw)

                # Infer session type
                session_type = _infer_session_type(course_name_raw, credits)

                record = TimetableRecord(
                    batch_raw=current_batch_raw,
                    program=current_batch_parsed.get("program", ""),
                    branch=current_batch_parsed.get("branch", ""),
                    semester=current_batch_parsed.get("semester", 0),
                    year=current_batch_parsed.get("year", 0),
                    day=day_name,
                    day_of_week=DAY_OF_WEEK[day_name],
                    start_time=slot["start"],
                    end_time=slot["end"],
                    course_code=course_code,
                    course_name=course_name_clean,
                    section=section,
                    credits=credits,
                    course_type=course_type,
                    session_type=session_type,
                    faculty_initials=faculty,
                    room=room,
                )
                records.append(record)

    logger.info(
        "Parsed %d class records across %d time slots",
        len(records), len(time_slots),
    )

    wb.close()
    return records


def _parse_batch_v2(batch_str: str) -> dict:
    """Parse Autumn 2026 batch strings.

    Examples:
        'Btech 1st Yr'        -> program=BTech, year=1, semester=1
        'Btech 2nd Year'      -> program=BTech, year=2, semester=3
        'Btech 3rd Yr (Core)' -> program=BTech, year=3, semester=5
        'BS-MS (IT)'          -> program=BS-MS, branch=IT, year=1, semester=1
        'BS-MS (DS + AI)'     -> program=BS-MS, branch=DS + AI, year=1, semester=1
        'MSc (AA)'            -> program=MSc, branch=AA, year=1, semester=1
        'MSc (IT) Core'       -> program=MSc, branch=IT, year=1, semester=1
        'MSc DS (Core)'       -> program=MSc, branch=DS, year=1, semester=1
        'Mtech (Core)'        -> program=MTech, year=1, semester=1
        'Elective'            -> program=Elective, year=0, semester=0
    """
    s = batch_str.strip()

    # Special case: Elective
    if s.lower() == "elective":
        return {"program": "Elective", "branch": "", "semester": 0, "year": 0}

    # BTech with year: "Btech 1st Yr", "Btech 2nd Year"
    m = BATCH_V2_YEAR_RE.search(s)
    if m:
        year = int(m.group(2))
        # Autumn semester is the odd one: year 1 -> sem 1, year 2 -> sem 3, etc.
        semester = (year - 1) * 2 + 1
        return {"program": "BTech", "branch": "", "semester": semester, "year": year}

    # BS-MS: "BS-MS (IT)", "BS-MS (DS + AI)"
    m2 = re.match(r"BS-MS\s*\(([^)]+)\)", s, re.IGNORECASE)
    if m2:
        return {"program": "BS-MS", "branch": m2.group(1).strip(), "semester": 1, "year": 1}

    # MSc with branch: "MSc (AA)", "MSc (IT) Core", "MSC (AA)"
    m3 = re.match(r"MSc\s*\(([^)]+)\)", s, re.IGNORECASE)
    if m3:
        return {"program": "MSc", "branch": m3.group(1).strip(), "semester": 1, "year": 1}

    # MSc DS: "MSc DS (Core)", "MSc DS Core"
    m4 = re.match(r"MSc\s+DS", s, re.IGNORECASE)
    if m4:
        return {"program": "MSc", "branch": "DS", "semester": 1, "year": 1}

    # MTech: "Mtech (Core)"
    m5 = re.match(r"MTech", s, re.IGNORECASE)
    if m5:
        return {"program": "MTech", "branch": "", "semester": 1, "year": 1}

    # "Btech Core" (no year number) -- assume it refers to a generic core slot
    m6 = BATCH_V2_CORE_RE.match(s)
    if m6:
        return {"program": "BTech", "branch": "", "semester": 0, "year": 0}

    # Fallback to V1 regex
    parsed = _parse_batch(s)
    if parsed["program"]:
        return parsed

    logger.warning("Could not parse V2 batch string: %r", s)
    return {"program": s, "branch": "", "semester": 0, "year": 0}


def _extract_section_from_code(course_code: str) -> tuple[str, Optional[str]]:
    """Extract section from course code in the V2 format.

    >>> _extract_section_from_code('IC101 (A)')
    ('IC101', 'A')
    >>> _extract_section_from_code('IC104 ( C )')
    ('IC104', 'C')
    >>> _extract_section_from_code('PC1 (ICT)')
    ('PC1 (ICT)', None)  # Not a section, it's a branch qualifier
    >>> _extract_section_from_code('ED312')
    ('ED312', None)
    """
    m = CODE_SECTION_RE.search(course_code)
    if m:
        sec = m.group(1).strip().upper()
        # Only treat single letters A-Z or digits 1-9 as sections
        # Multi-char like 'ICT', 'CS' are branch qualifiers, not sections
        if len(sec) == 1 and (sec.isalpha() or sec.isdigit()):
            clean_code = course_code[:m.start()].strip()
            return clean_code, sec
    return course_code.strip(), None


def _parse_v2_3col(
    xlsx_path: str | Path,
    sheet_name: str,
) -> list[TimetableRecord]:
    """Parse the Autumn 2026 format: 3 columns per day (code, faculty, room),
    batch in C3, section embedded in course code."""
    import openpyxl

    xlsx_path = Path(xlsx_path)
    logger.info("Opening timetable (V2 3-col format): %s", xlsx_path)
    wb = openpyxl.load_workbook(str(xlsx_path), data_only=True)
    ws = wb[sheet_name]
    max_row = ws.max_row
    logger.info("Sheet '%s': %d rows x %d cols", sheet_name, max_row, ws.max_column)

    # Step 1: Discover time-slot boundaries
    time_slots: list[dict] = []
    for row_idx in range(1, max_row + 1):
        a_val = _cell_str(ws, row_idx, 1)
        m = TIME_SLOT_RE.match(a_val)
        if m:
            raw_start = m.group(1)
            raw_end = m.group(2)
            # Handle PM times: "2:00 - 2:50PM" -> 14:00 - 14:50
            is_pm = "pm" in a_val.lower()
            start_t = _zero_pad_time(raw_start)
            end_t = _zero_pad_time(raw_end)
            if is_pm or (int(raw_start.split(":")[0]) < 6):
                sh, sm = start_t.split(":")
                eh, em = end_t.split(":")
                sh_i, eh_i = int(sh), int(eh)
                if sh_i < 12 and sh_i != 0:
                    start_t = "{:02d}:{:02d}".format(sh_i + 12, int(sm))
                if eh_i < 12 and eh_i != 0:
                    end_t = "{:02d}:{:02d}".format(eh_i + 12, int(em))
            time_slots.append({
                "row": row_idx,
                "start": start_t,
                "end": end_t,
                "raw": a_val,
            })

    if not time_slots:
        raise ValueError("No time-slot rows found in column A.")

    logger.info("Found %d time slots: %s", len(time_slots), [s["raw"] for s in time_slots])

    # Step 2: Parse each time-slot block
    records: list[TimetableRecord] = []

    for slot_idx, slot in enumerate(time_slots):
        slot_start_row = slot["row"]
        if slot_idx + 1 < len(time_slots):
            slot_end_row = time_slots[slot_idx + 1]["row"] - 1
        else:
            slot_end_row = max_row

        current_batch_raw = ""
        current_batch_parsed: dict = {}

        for row_idx in range(slot_start_row, slot_end_row + 1):
            # Check C3 for batch header
            c3 = _cell_str(ws, row_idx, 3)
            if c3:
                current_batch_raw = c3
                current_batch_parsed = _parse_batch_v2(c3)

            # Skip rows with no batch context or header rows
            if not current_batch_raw:
                continue

            # Skip column-header rows ("Course", "Name", "Room")
            c4 = _cell_str(ws, row_idx, 4)
            if c4.lower() in ("course", ""):
                # Check if any other day column has data
                has_any_data = False
                for day_name_inner, cols in DAY_COLUMNS_V2.items():
                    if _cell_str(ws, row_idx, cols[0]) and _cell_str(ws, row_idx, cols[0]).lower() != "course":
                        has_any_data = True
                        break
                if not has_any_data:
                    continue

            # Determine course_type based on batch
            is_elective = current_batch_parsed.get("program", "") == "Elective"
            course_type = "Elective" if is_elective else "Core"

            # Extract classes for each day
            for day_name, cols in DAY_COLUMNS_V2.items():
                code_col, faculty_col, room_col = cols

                course_code_raw = _cell_str(ws, row_idx, code_col)
                faculty = _cell_str(ws, row_idx, faculty_col)
                room = _cell_str(ws, row_idx, room_col)

                if not course_code_raw or course_code_raw.lower() in ("course", "."):
                    continue

                # Extract section from code: "IC101 (A)" -> code="IC101", section="A"
                course_code, section = _extract_section_from_code(course_code_raw)

                # For course name, use the code as the name (V2 doesn't have
                # a separate name column -- the "Name" column is actually faculty)
                # We'll use just the course code as course_name for now
                course_name = course_code

                # Infer session type from name
                session_type = "lab" if "lab" in course_code.lower() else "lecture"

                record = TimetableRecord(
                    batch_raw=current_batch_raw,
                    program=current_batch_parsed.get("program", ""),
                    branch=current_batch_parsed.get("branch", ""),
                    semester=current_batch_parsed.get("semester", 0),
                    year=current_batch_parsed.get("year", 0),
                    day=day_name,
                    day_of_week=DAY_OF_WEEK[day_name],
                    start_time=slot["start"],
                    end_time=slot["end"],
                    course_code=course_code,
                    course_name=course_name,
                    section=section,
                    credits="",
                    course_type=course_type,
                    session_type=session_type,
                    faculty_initials=faculty,
                    room=room,
                )
                records.append(record)

    logger.info(
        "Parsed %d class records across %d time slots (V2 format)",
        len(records), len(time_slots),
    )
    wb.close()
    return records


# ── Convenience views ─────────────────────────────────────────────────────


def student_view(
    records: list[TimetableRecord],
    program: str,
    semester: int,
    branch: str = "",
    section: Optional[str] = None,
) -> list[TimetableRecord]:
    """
    Filter records for a specific student's batch.

    Parameters
    ----------
    program : str
        e.g. "BTech"
    semester : int
        e.g. 2
    branch : str
        e.g. "ICT + CS", "EVD", "MnC". Empty string matches all.
    section : str or None
        e.g. "A". If provided, only returns courses for that section
        (plus courses with no section, which are common to all).
    """
    filtered = []
    for r in records:
        # Match program and semester
        if r.program.lower() != program.lower():
            continue
        if r.semester != semester:
            continue

        # Match branch (if specified)
        if branch:
            # Handle combined branches like "ICT + CS" — both ICT and CS students see these
            r_branch = r.branch.lower()
            q_branch = branch.lower()
            if q_branch not in r_branch and r_branch not in q_branch:
                # Check if the branch field contains the query
                # e.g. "ICT + CS" contains "ICT" and "CS"
                branch_parts = [b.strip().lower() for b in r.branch.split("+")]
                if q_branch not in branch_parts and not any(q_branch in bp for bp in branch_parts):
                    # Also check for patterns like "CS-Only" matching "CS"
                    if not (q_branch.replace("-only", "") in r_branch or r_branch.replace("-only", "") in q_branch):
                        continue

        # Match section (if specified)
        if section:
            # Include courses with no section (common to all) and matching section
            if r.section is not None and r.section.upper() != section.upper():
                continue

        filtered.append(r)

    return sorted(filtered, key=lambda r: (r.day_of_week, r.start_time))


def faculty_view(
    records: list[TimetableRecord],
    faculty_initials: str,
) -> list[TimetableRecord]:
    """
    Filter records for a specific faculty member's teaching schedule.

    Matches against faculty_initials (case-insensitive). Also handles
    combined faculty like "SS/VS" — will match if the query is "SS" or "VS".
    """
    query = faculty_initials.strip().upper()
    filtered = []
    for r in records:
        initials_upper = r.faculty_initials.upper()
        # Exact match
        if initials_upper == query:
            filtered.append(r)
            continue
        # Handle combined faculty: "SS/VS" matches "SS" or "VS"
        if "/" in initials_upper:
            parts = [p.strip() for p in initials_upper.split("/")]
            if query in parts:
                filtered.append(r)

    return sorted(filtered, key=lambda r: (r.day_of_week, r.start_time))


def to_dicts(records: list[TimetableRecord]) -> list[dict]:
    """Convert records to plain dicts (for JSON serialization)."""
    return [asdict(r) for r in records]


# ── Statistics ────────────────────────────────────────────────────────────


def summary(records: list[TimetableRecord]) -> dict:
    """Generate summary statistics from parsed records."""
    batches = set()
    courses = set()
    faculty = set()
    sections = set()

    for r in records:
        batches.add(r.batch_raw)
        courses.add(r.course_code)
        faculty.add(r.faculty_initials)
        if r.section:
            sections.add(r.section)

    return {
        "total_records": len(records),
        "unique_batches": len(batches),
        "unique_courses": len(courses),
        "unique_faculty": len(faculty),
        "unique_sections": sorted(sections),
        "batches": sorted(batches),
        "session_types": {
            "lecture": sum(1 for r in records if r.session_type == "lecture"),
            "lab": sum(1 for r in records if r.session_type == "lab"),
            "tutorial": sum(1 for r in records if r.session_type == "tutorial"),
        },
    }


# ── CLI ───────────────────────────────────────────────────────────────────


def _cli():
    """Command-line interface for the timetable parser."""
    import argparse

    parser = argparse.ArgumentParser(
        description="Parse DAU Lecture Time-Table XLSX into structured records."
    )
    parser.add_argument("xlsx_path", help="Path to the XLSX timetable file.")
    parser.add_argument("--sheet", default="Time-Table", help="Sheet name (default: Time-Table).")
    parser.add_argument("--json", dest="json_out", metavar="PATH", help="Write parsed records as JSON to PATH.")
    parser.add_argument("--csv", dest="csv_out", metavar="PATH", help="Write parsed records as CSV to PATH.")
    parser.add_argument("--summary", action="store_true", help="Print summary statistics.")
    parser.add_argument(
        "--student", nargs="+", metavar=("PROGRAM", "SEM"),
        help="Filter student view: --student BTech 2 'ICT + CS' A"
    )
    parser.add_argument(
        "--faculty", metavar="INITIALS",
        help="Filter faculty view: --faculty PD"
    )

    args = parser.parse_args()

    logging.basicConfig(level=logging.INFO, format="%(levelname)s: %(message)s")

    records = parse_timetable(args.xlsx_path, sheet_name=args.sheet)

    # Apply filters
    if args.student:
        program = args.student[0]
        sem = int(args.student[1])
        branch = args.student[2] if len(args.student) > 2 else ""
        section = args.student[3] if len(args.student) > 3 else None
        records = student_view(records, program, sem, branch, section)
        print(f"\n[Student Timetable] {program} Sem-{sem} {branch} {f'Section {section}' if section else ''}")

    if args.faculty:
        records = faculty_view(records, args.faculty)
        print(f"\n[Faculty Timetable] {args.faculty}")

    # Output
    if args.summary:
        stats = summary(records)
        print("\n[Summary]")
        for k, v in stats.items():
            if isinstance(v, list) and len(v) > 10:
                print(f"  {k}: [{len(v)} items]")
            else:
                print(f"  {k}: {v}")

    if args.json_out:
        with open(args.json_out, "w", encoding="utf-8") as f:
            json.dump(to_dicts(records), f, indent=2, ensure_ascii=False)
        print(f"\nJSON written to {args.json_out} ({len(records)} records)")

    if args.csv_out:
        import csv
        fieldnames = list(asdict(records[0]).keys()) if records else []
        with open(args.csv_out, "w", encoding="utf-8", newline="") as f:
            writer = csv.DictWriter(f, fieldnames=fieldnames)
            writer.writeheader()
            writer.writerows(to_dicts(records))
        print(f"\nCSV written to {args.csv_out} ({len(records)} records)")

    if not (args.json_out or args.csv_out or args.summary or args.student or args.faculty):
        # Default: print all records as a formatted table
        print(f"\nParsed {len(records)} class records.\n")
        print(f"{'Batch':<35} {'Day':<10} {'Time':<14} {'Code':<10} {'Course':<40} {'Sec':<5} {'Type':<8} {'Faculty':<8} {'Room':<10}")
        print("─" * 145)
        for r in records[:50]:
            sec_display = r.section or "—"
            print(f"{r.batch_raw:<35} {r.day:<10} {r.start_time}-{r.end_time}  {r.course_code:<10} {r.course_name:<40} {sec_display:<5} {r.session_type:<8} {r.faculty_initials:<8} {r.room:<10}")
        if len(records) > 50:
            print(f"\n  ... and {len(records) - 50} more records. Use --json or --csv to export all.")


if __name__ == "__main__":
    _cli()
