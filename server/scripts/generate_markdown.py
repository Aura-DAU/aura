import openpyxl
import glob
import os
import re
import json

def get_24hr_time(t_str):
    if not t_str: return ""
    parts = str(t_str).split('-')
    if len(parts) != 2: return t_str
    
    def norm(s):
        s = s.strip().lower()
        if not s: return s
        is_pm = "pm" in s
        is_am = "am" in s
        s = s.replace("am", "").replace("pm", "").strip()
        p = s.split(':')
        if len(p) == 2:
            try:
                hr = int(p[0])
                if is_pm and hr != 12: hr += 12
                elif is_am and hr == 12: hr = 0
                elif not is_pm and not is_am:
                    # DAU heuristic
                    if 1 <= hr <= 7: hr += 12
                return f"{hr:02d}:{int(p[1]):02d}"
            except:
                pass
        return s
        
    return f"{norm(parts[0])}-{norm(parts[1])}"

def get_merged_value(sheet, row, col):
    cell = sheet.cell(row=row, column=col)
    for merged_range in sheet.merged_cells.ranges:
        if cell.coordinate in merged_range:
            top_left = sheet.cell(row=merged_range.min_row, column=merged_range.min_col)
            return top_left.value
    return cell.value

def main():
    base_dir = os.path.dirname(os.path.dirname(os.path.dirname(__file__)))
    xl_path = os.path.join(base_dir, "Lecture_TT_Autumn2026-27_v9 (1).xlsx")
    out_dir = os.path.join(base_dir, "data", "academics", "timetable")
    
    if not os.path.exists(out_dir):
        os.makedirs(out_dir)
        
    # Clear existing markdown files
    for f in glob.glob(os.path.join(out_dir, "*.md")):
        os.remove(f)
        
    print(f"Loading {xl_path}...")
    wb = openpyxl.load_workbook(xl_path, data_only=True)
    
    # Map faculty
    fac_sheet = wb['FacultyNameAbbreviations']
    faculty_map = {}
    for row in range(2, fac_sheet.max_row + 1):
        full_name = fac_sheet.cell(row=row, column=1).value
        short_name = fac_sheet.cell(row=row, column=2).value
        if full_name and short_name:
            faculty_map[str(short_name).strip()] = str(full_name).strip()
            
    # Parse timetable
    sheet = wb['Lecture (Update)']
    days_cols = {
        'Monday': 4,
        'Tuesday': 8,
        'Wednesday': 12,
        'Thursday': 16,
        'Friday': 20
    }
    
    slots_by_batch = {}
    raw_records = []
    
    for row in range(6, sheet.max_row + 1):
        time_val = get_merged_value(sheet, row, 1)
        if not time_val: continue
        time_str = str(time_val).strip()
        if ':' not in time_str: continue
        
        time_24 = get_24hr_time(time_str)
        t_start = time_24.split('-')[0] if '-' in time_24 else time_24
        t_end = time_24.split('-')[1] if '-' in time_24 else ""
        
        batch_val = get_merged_value(sheet, row, 3)
        batch = str(batch_val).strip() if batch_val else "Unknown"
        # Normalize batch name (fix typos from excel)
        if batch == "Btech 3r Yr": batch = "Btech 3rd Yr"
        if batch == "Btech 3rd Yr (Core)": batch = "Btech 3rd Yr"
        if batch == "Btech 2nd Year": batch = "Btech 2nd Yr"
        if batch == "MSC (AA)": batch = "MSc (AA)"
        if batch == "MSc (IT) Core": batch = "MSc IT (Core)"
        if batch == "MSc (DS) Core": batch = "MSc DS (Core)"
        
        is_elective = "elective" in batch.lower()
        if is_elective:
            batch = "Electives_All"
            # we also map the batch field in JSON to 'Elective' so BATCH_MAP finds it
            json_batch = "Elective" 
        else:
            json_batch = batch
            
        if batch not in slots_by_batch:
            slots_by_batch[batch] = []
            
        for day, col in days_cols.items():
            course_val = get_merged_value(sheet, row, col)
            if not course_val: continue
            course = str(course_val).strip().replace('\n', ' ')
            if not course: continue
            
            # The JSON needs 'course_code'. e.g. "IC101 (A) - Intro to ICT" -> "IC101 (A)"
            ccode = course
            
            faculty_val = get_merged_value(sheet, row, col+1)
            faculty_code = str(faculty_val).strip().replace('\n', ' ') if faculty_val else ""
            faculty_full = faculty_map.get(faculty_code, faculty_code)
            
            room_val = get_merged_value(sheet, row, col+2)
            room = str(room_val).strip().replace('\n', ' ') if room_val else "TBA"
            if not room: room = "TBA"
            
            slots_by_batch[batch].append({
                'day': day,
                'time': time_24,
                'course': course,
                'faculty': faculty_full,
                'room': room
            })
            
            # Also append to raw_records
            raw_records.append({
                "batch": json_batch,
                "day": day,
                "course_code": ccode,
                "start": t_start,
                "end": t_end,
                "faculty": faculty_full,
                "room": room,
                "is_elective": is_elective
            })

    # Write out markdowns
    for batch, slots in slots_by_batch.items():
        if not slots: continue
        
        safe_name = batch.replace(" ", "_").replace("(", "").replace(")", "").replace("+", "_").replace("__", "_")
        if batch == "Unknown" and len(slots) > 0:
            safe_name = "Btech_Core_year_unspecified_in_source_verify_manually"
            
        filepath = os.path.join(out_dir, f"{safe_name}.md")
        with open(filepath, 'w', encoding='utf-8') as f:
            f.write(f"# Timetable — {batch}\n\n")
            f.write("Source: official DAU Autumn 2026-27 lecture timetable.\n\n")
            f.write("| Day | Time | Course | Faculty | Room |\n")
            f.write("|---|---|---|---|---|\n")
            
            # Sort slots by day then time
            day_order = {'Monday':0, 'Tuesday':1, 'Wednesday':2, 'Thursday':3, 'Friday':4}
            slots.sort(key=lambda x: (day_order.get(x['day'], 99), x['time']))
            
            for s in slots:
                f.write(f"| {s['day']} | {s['time']} | {s['course']} | {s['faculty']} | {s['room']} |\n")
                
    print(f"Generated {len(slots_by_batch)} markdown files in {out_dir}.")
    
    # Write JSON
    json_path = os.path.join(out_dir, "_raw_records.json")
    with open(json_path, "w", encoding="utf-8") as jf:
        json.dump(raw_records, jf, indent=2)
    print(f"Generated _raw_records.json with {len(raw_records)} records.")

if __name__ == "__main__":
    main()
