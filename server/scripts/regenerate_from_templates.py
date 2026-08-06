import os
import glob
import re
import json
import openpyxl

def get_merged_value(sheet, row, col):
    cell = sheet.cell(row=row, column=col)
    for merged_range in sheet.merged_cells.ranges:
        if cell.coordinate in merged_range:
            top_left = sheet.cell(row=merged_range.min_row, column=merged_range.min_col)
            return top_left.value
    return cell.value

def to_24h(time_str):
    time_str = str(time_str).strip()
    if not time_str: return time_str
    
    parts = time_str.split('-')
    if len(parts) != 2: return time_str
    
    def norm(s):
        s = s.strip().lower()
        if not s: return s
        is_pm = "pm" in s
        is_am = "am" in s
        s = s.replace("am", "").replace("pm", "").strip()
        p = s.split(':')
        if len(p) == 2:
            try:
                hr = int(p[0])
                if is_pm and hr != 12: hr += 12
                elif is_am and hr == 12: hr = 0
                elif not is_pm and not is_am:
                    if 1 <= hr <= 7: hr += 12
                return f"{hr:02d}:{int(p[1]):02d}"
            except:
                pass
        return s
        
    return f"{norm(parts[0])}-{norm(parts[1])}"

def get_base_code(s):
    s = s.upper()
    code = s.split()[0]
    sec = ""
    if "(SEC " in s: sec = s.split("(SEC ")[1][0]
    elif "(A" in s: sec = "A"
    elif "(B" in s: sec = "B"
    elif "(C" in s: sec = "C"
    elif "(D" in s: sec = "D"
    return code, sec

def main():
    base_dir = os.path.dirname(os.path.dirname(os.path.dirname(__file__)))
    xl_path = os.path.join(base_dir, "Lecture_TT_Autumn2026-27_v9 (1).xlsx")
    md_dir = os.path.join(base_dir, "data", "academics", "timetable")
    
    print("Loading Excel...")
    wb = openpyxl.load_workbook(xl_path, data_only=True)
    sheet = wb['Lecture (Update)']
    days_cols = {'Monday': 4, 'Tuesday': 8, 'Wednesday': 12, 'Thursday': 16, 'Friday': 20}
    
    excel_lookup = {}
    json_records = []
    
    for row in range(6, sheet.max_row + 1):
        time_val = get_merged_value(sheet, row, 1)
        if not time_val or ':' not in str(time_val): continue
        time_24 = to_24h(str(time_val).strip())
        t_start = time_24.split('-')[0] if '-' in time_24 else time_24
        t_end = time_24.split('-')[1] if '-' in time_24 else ""
        
        batch_val = get_merged_value(sheet, row, 3)
        batch = str(batch_val).strip() if batch_val else "Unknown"
        if batch == "Btech 3r Yr": batch = "Btech 3rd Yr"
        if batch == "Btech 3rd Yr (Core)": batch = "Btech 3rd Yr"
        if batch == "Btech 2nd Year": batch = "Btech 2nd Yr"
        if batch == "MSC (AA)": batch = "MSc (AA)"
        if batch == "MSc (IT) Core": batch = "MSc IT (Core)"
        if batch == "MSc (DS) Core": batch = "MSc DS (Core)"
        
        is_elective = "elective" in batch.lower()
        if is_elective: json_batch = "Elective"
        else: json_batch = batch
        
        for day, col in days_cols.items():
            course_val = get_merged_value(sheet, row, col)
            if not course_val: continue
            
            c_str = str(course_val).strip().replace('\n', ' ')
            if not c_str: continue
            
            fac_val = get_merged_value(sheet, row, col+1)
            room_val = get_merged_value(sheet, row, col+2)
            
            c_code, c_sec = get_base_code(c_str)
            key = (day, c_code, c_sec)
            
            # Save for JSON
            json_records.append({
                "batch": json_batch,
                "day": day,
                "start": t_start,
                "end": t_end,
                "course_code": c_code,
                "faculty": str(fac_val).strip() if fac_val else "",
                "room": str(room_val).strip() if room_val else "",
                "is_elective": is_elective
            })
            
            # Save for MD lookup
            excel_lookup[key] = {
                "time_24": time_24,
                "faculty": str(fac_val).strip() if fac_val else "",
                "room": str(room_val).strip() if room_val else ""
            }
            
    md_files = glob.glob(os.path.join(md_dir, "*.md"))
    rebuilt_count = 0
    
    for md_file in md_files:
        with open(md_file, 'r', encoding='utf-8') as f:
            lines = f.readlines()
            
        new_lines = []
        for line in lines:
            if line.startswith('|') and not line.startswith('| Day |') and not line.startswith('|---|'):
                parts = [p.strip() for p in line.split('|')]
                if len(parts) >= 6:
                    day = parts[1]
                    course_str_md = parts[3]
                    
                    c_code, c_sec = get_base_code(course_str_md)
                    key = (day, c_code, c_sec)
                    
                    if key in excel_lookup:
                        ex_data = excel_lookup[key]
                        # Overwrite with pristine time, faculty, room
                        new_line = f"| {day} | {ex_data['time_24']} | {course_str_md} | {ex_data['faculty']} | {ex_data['room']} |\n"
                        new_lines.append(new_line)
                    else:
                        new_lines.append(line)
                else:
                    new_lines.append(line)
            else:
                new_lines.append(line)
                
        with open(md_file, 'w', encoding='utf-8') as f:
            f.writelines(new_lines)
        rebuilt_count += 1
        
    json_path = os.path.join(md_dir, "_raw_records.json")
    with open(json_path, 'w', encoding='utf-8') as f:
        json.dump(json_records, f, indent=2)
        
    print(f"Successfully rebuilt {rebuilt_count} markdown files!")
    print(f"Generated {len(json_records)} accurate JSON records for the database.")

if __name__ == "__main__":
    main()
